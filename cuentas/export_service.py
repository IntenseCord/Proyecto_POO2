"""
Servicio de exportación de reportes contables a PDF y Excel
"""
from decimal import Decimal
from datetime import datetime

# Constantes reutilizables
HEADER_CODIGO = 'Código'
CURRENCY_FORMAT = '$#,##0.00'


class ExportadorReportes:
    """Clase base para exportación de reportes"""
    
    def __init__(self, reporte_data):
        self.reporte_data = reporte_data
    
    def exportar_pdf(self):
        """Exporta el reporte a PDF"""
        raise NotImplementedError("Debe ser implementado por las subclases")
    
    def exportar_excel(self):
        """Exporta el reporte a Excel"""
        raise NotImplementedError("Debe ser implementado por las subclases")


class ExportadorBalanceGeneral(ExportadorReportes):
    """Exportador específico para Balance General"""
    
    def exportar_pdf(self):
        """
        Exporta el Balance General a PDF usando ReportLab
        """
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
            from io import BytesIO
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, 
                                  topMargin=72, bottomMargin=18)
            
            # Container for the 'Flowable' objects
            elements = []
            styles = getSampleStyleSheet()
            
            # Título del reporte
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#2c3e50'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            
            empresa = self.reporte_data.get('empresa')
            fecha_inicio = self.reporte_data.get('fecha_inicio')
            fecha_fin = self.reporte_data.get('fecha_fin')
            
            # Encabezado
            elements.append(Paragraph(f"<b>{empresa.nombre}</b>", title_style))
            elements.append(Paragraph("Balance General", title_style))
            
            # Período
            if fecha_inicio and fecha_fin:
                periodo_text = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
            elif fecha_fin:
                periodo_text = f"Al {fecha_fin.strftime('%d/%m/%Y')}"
            else:
                periodo_text = f"Al {datetime.now().strftime('%d/%m/%Y')}"
            
            elements.append(Paragraph(periodo_text, styles['Normal']))
            elements.append(Spacer(1, 20))
            
            # ACTIVOS
            elements.append(Paragraph("<b>ACTIVOS</b>", styles['Heading2']))
            activos_data = [[HEADER_CODIGO, 'Cuenta', 'Monto']]
            for activo in self.reporte_data.get('activos', []):
                activos_data.append([
                    activo['codigo'],
                    activo['nombre'],
                    f"${activo['monto']:,.2f}"
                ])
            
            activos_data.append(['', 'TOTAL ACTIVOS', 
                               f"${self.reporte_data['totales']['activos']:,.2f}"])
            
            activos_table = Table(activos_data, colWidths=[1*inch, 4*inch, 1.5*inch])
            activos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(activos_table)
            elements.append(Spacer(1, 20))
            
            # PASIVOS
            elements.append(Paragraph("<b>PASIVOS</b>", styles['Heading2']))
            pasivos_data = [[HEADER_CODIGO, 'Cuenta', 'Monto']]
            for pasivo in self.reporte_data.get('pasivos', []):
                pasivos_data.append([
                    pasivo['codigo'],
                    pasivo['nombre'],
                    f"${pasivo['monto']:,.2f}"
                ])
            
            pasivos_data.append(['', 'TOTAL PASIVOS', 
                               f"${self.reporte_data['totales']['pasivos']:,.2f}"])
            
            pasivos_table = Table(pasivos_data, colWidths=[1*inch, 4*inch, 1.5*inch])
            pasivos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(pasivos_table)
            elements.append(Spacer(1, 20))
            
            # PATRIMONIO
            elements.append(Paragraph("<b>PATRIMONIO</b>", styles['Heading2']))
            patrimonio_data = [[HEADER_CODIGO, 'Cuenta', 'Monto']]
            for patrimonio in self.reporte_data.get('patrimonios', []):
                patrimonio_data.append([
                    patrimonio['codigo'],
                    patrimonio['nombre'],
                    f"${patrimonio['monto']:,.2f}"
                ])
            
            # Agregar utilidad del período
            utilidad_periodo = self.reporte_data.get('utilidad_periodo', 0)
            if utilidad_periodo != 0:
                label = 'Utilidad del Período' if utilidad_periodo > 0 else 'Pérdida del Período'
                patrimonio_data.append(['', label, f"${utilidad_periodo:,.2f}"])
            
            patrimonio_data.append(['', 'TOTAL PATRIMONIO', 
                                  f"${self.reporte_data['totales']['patrimonio_con_utilidad']:,.2f}"])
            
            patrimonio_table = Table(patrimonio_data, colWidths=[1*inch, 4*inch, 1.5*inch])
            patrimonio_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(patrimonio_table)
            elements.append(Spacer(1, 20))
            
            # Ecuación contable
            ecuacion_data = [
                ['ACTIVOS', 'PASIVOS + PATRIMONIO', 'DIFERENCIA'],
                [
                    f"${self.reporte_data['totales']['activos']:,.2f}",
                    f"${self.reporte_data['totales']['pasivo_patrimonio']:,.2f}",
                    f"${self.reporte_data['diferencia']:,.2f}"
                ]
            ]
            
            ecuacion_table = Table(ecuacion_data, colWidths=[2*inch, 2*inch, 2*inch])
            ecuacion_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('FONTSIZE', (0, 1), (-1, 1), 14),
                ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 1), (-1, 1), 12),
                ('BOTTOMPADDING', (0, 1), (-1, 1), 12),
                ('GRID', (0, 0), (-1, -1), 2, colors.black)
            ]))
            elements.append(ecuacion_table)
            
            # Build PDF
            doc.build(elements)
            
            # FileResponse sets the Content-Disposition header so that browsers
            # present the option to save the file.
            buffer.seek(0)
            return buffer
            
        except ImportError:
            raise ImportError("ReportLab no está instalado. Ejecuta: pip install reportlab")
    
    def exportar_excel(self):
        """
        Exporta el Balance General a Excel usando openpyxl
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from io import BytesIO
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Balance General"
            
            # Estilos
            title_font = Font(name='Arial', size=16, bold=True)
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            empresa = self.reporte_data.get('empresa')
            fecha_inicio = self.reporte_data.get('fecha_inicio')
            fecha_fin = self.reporte_data.get('fecha_fin')
            
            # Título
            ws['A1'] = empresa.nombre
            ws['A1'].font = title_font
            ws['A2'] = 'Balance General'
            ws['A2'].font = Font(name='Arial', size=14, bold=True)
            
            # Período
            if fecha_inicio and fecha_fin:
                periodo_text = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
            elif fecha_fin:
                periodo_text = f"Al {fecha_fin.strftime('%d/%m/%Y')}"
            else:
                periodo_text = f"Al {datetime.now().strftime('%d/%m/%Y')}"
            ws['A3'] = periodo_text
            
            row = 5
            
            # ACTIVOS
            ws[f'A{row}'] = 'ACTIVOS'
            ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True)
            row += 1
            
            ws[f'A{row}'] = HEADER_CODIGO
            ws[f'B{row}'] = 'Cuenta'
            ws[f'C{row}'] = 'Monto'
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].font = header_font
                ws[f'{col}{row}'].fill = header_fill
                ws[f'{col}{row}'].border = border
            row += 1
            
            for activo in self.reporte_data.get('activos', []):
                ws[f'A{row}'] = activo['codigo']
                ws[f'B{row}'] = activo['nombre']
                ws[f'C{row}'] = float(activo['monto'])
                ws[f'C{row}'].number_format = CURRENCY_FORMAT
                row += 1
            
            ws[f'B{row}'] = 'TOTAL ACTIVOS'
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'C{row}'] = float(self.reporte_data['totales']['activos'])
            ws[f'C{row}'].number_format = CURRENCY_FORMAT
            ws[f'C{row}'].font = Font(bold=True)
            ws[f'C{row}'].fill = total_fill
            row += 2
            
            # PASIVOS
            ws[f'A{row}'] = 'PASIVOS'
            ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True)
            row += 1
            
            ws[f'A{row}'] = HEADER_CODIGO
            ws[f'B{row}'] = 'Cuenta'
            ws[f'C{row}'] = 'Monto'
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].font = header_font
                ws[f'{col}{row}'].fill = header_fill
                ws[f'{col}{row}'].border = border
            row += 1
            
            for pasivo in self.reporte_data.get('pasivos', []):
                ws[f'A{row}'] = pasivo['codigo']
                ws[f'B{row}'] = pasivo['nombre']
                ws[f'C{row}'] = float(pasivo['monto'])
                ws[f'C{row}'].number_format = CURRENCY_FORMAT
                row += 1
            
            ws[f'B{row}'] = 'TOTAL PASIVOS'
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'C{row}'] = float(self.reporte_data['totales']['pasivos'])
            ws[f'C{row}'].number_format = CURRENCY_FORMAT
            ws[f'C{row}'].font = Font(bold=True)
            ws[f'C{row}'].fill = total_fill
            row += 2
            
            # PATRIMONIO
            ws[f'A{row}'] = 'PATRIMONIO'
            ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True)
            row += 1
            
            ws[f'A{row}'] = HEADER_CODIGO
            ws[f'B{row}'] = 'Cuenta'
            ws[f'C{row}'] = 'Monto'
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].font = header_font
                ws[f'{col}{row}'].fill = header_fill
                ws[f'{col}{row}'].border = border
            row += 1
            
            for patrimonio in self.reporte_data.get('patrimonios', []):
                ws[f'A{row}'] = patrimonio['codigo']
                ws[f'B{row}'] = patrimonio['nombre']
                ws[f'C{row}'] = float(patrimonio['monto'])
                ws[f'C{row}'].number_format = CURRENCY_FORMAT
                row += 1
            
            # Utilidad del período
            utilidad_periodo = self.reporte_data.get('utilidad_periodo', 0)
            if utilidad_periodo != 0:
                label = 'Utilidad del Período' if utilidad_periodo > 0 else 'Pérdida del Período'
                ws[f'B{row}'] = label
                ws[f'C{row}'] = float(utilidad_periodo)
                ws[f'C{row}'].number_format = CURRENCY_FORMAT
                row += 1
            
            ws[f'B{row}'] = 'TOTAL PATRIMONIO'
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'C{row}'] = float(self.reporte_data['totales']['patrimonio_con_utilidad'])
            ws[f'C{row}'].number_format = CURRENCY_FORMAT
            ws[f'C{row}'].font = Font(bold=True)
            ws[f'C{row}'].fill = total_fill
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 20
            
            # Guardar en buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer
            
        except ImportError:
            raise ImportError("openpyxl no está instalado. Ejecuta: pip install openpyxl")

