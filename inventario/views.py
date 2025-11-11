from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Sum, F
from django.db import models
from django.views.decorators.http import require_http_methods, require_GET, require_POST
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import csrf_protect
from .models import Producto, Categoria, MovimientoInventario
from .forms import ProductoForm, CategoriaForm, MovimientoInventarioForm, ImportarProductosForm
from openpyxl import load_workbook
from io import BytesIO
from decimal import Decimal

# Constantes para evitar duplicación
DETALLE_PRODUCTO_URL = 'inventario:detalle_producto'
TEMPLATE_IMPORTAR_PRODUCTOS = 'inventario/importar_productos.html'

def _normalizar_valores_fila(row, indice, estados_validos):
    """
    Extrae y normaliza datos de una fila de Excel según el índice de encabezados.
    Devuelve un dict con los campos esperados por la importación.
    """
    obtener = lambda campo: row[indice[campo]].value if campo in indice else None

    codigo = obtener('codigo')
    if isinstance(codigo, str):
        codigo = codigo.strip()

    nombre = obtener('nombre')
    descripcion = obtener('descripcion') or ''
    categoria_nombre = obtener('categoria')

    # Normalizar cantidad
    try:
        cantidad = int(obtener('cantidad') or 0)
    except Exception:
        cantidad = 0

    # Normalizar precios
    try:
        precio_unitario = Decimal(str(obtener('precio_unitario') or '0'))
    except Exception:
        precio_unitario = Decimal('0.00')
    try:
        precio_venta_val = obtener('precio_venta')
        precio_venta = Decimal(str(precio_venta_val if precio_venta_val is not None else precio_unitario))
    except Exception:
        precio_venta = Decimal('0.00')

    # Normalizar stock mínimo
    try:
        stock_minimo = int(obtener('stock_minimo') or 5)
    except Exception:
        stock_minimo = 5

    # Normalizar estado
    estado_val = str(obtener('estado') or 'activo').strip().lower()
    estado = estado_val if estado_val in estados_validos else 'activo'

    return {
        'codigo': codigo,
        'nombre': nombre,
        'descripcion': descripcion,
        'categoria_nombre': categoria_nombre,
        'cantidad': cantidad,
        'precio_unitario': precio_unitario,
        'precio_venta': precio_venta,
        'stock_minimo': stock_minimo,
        'estado': estado,
    }

def _resolver_categoria_por_nombre(nombre_categoria, crear_categorias):
    if not nombre_categoria:
        return None
    nombre = str(nombre_categoria).strip()
    if not nombre:
        return None
    existente = Categoria.objects.filter(nombre__iexact=nombre).first()
    if existente:
        return existente
    return Categoria.objects.create(nombre=nombre) if crear_categorias else None

def _manejar_producto(data, user, crear_categorias):
    """
    Crea o actualiza un producto a partir de data normalizada.
    Retorna una tupla (creados, actualizados) con contadores 0/1.
    """
    categoria_obj = _resolver_categoria_por_nombre(data['categoria_nombre'], crear_categorias)

    if data['codigo']:
        producto, creado = Producto.objects.get_or_create(
            codigo=str(data['codigo']).strip(),
            defaults={
                'nombre': data['nombre'],
                'descripcion': data['descripcion'],
                'categoria': categoria_obj,
                'cantidad': data['cantidad'],
                'precio_unitario': data['precio_unitario'],
                'precio_venta': data['precio_venta'],
                'stock_minimo': data['stock_minimo'],
                'estado': data['estado'],
                'usuario_creador': user,
            }
        )
        if not creado:
            producto.nombre = data['nombre'] or producto.nombre
            producto.descripcion = data['descripcion'] or producto.descripcion
            producto.categoria = categoria_obj or producto.categoria
            producto.cantidad = data['cantidad'] if data['cantidad'] is not None else producto.cantidad
            producto.precio_unitario = data['precio_unitario'] or producto.precio_unitario
            producto.precio_venta = data['precio_venta'] or producto.precio_venta
            producto.stock_minimo = data['stock_minimo'] if data['stock_minimo'] is not None else producto.stock_minimo
            producto.estado = data['estado'] or producto.estado
            producto.save()
            return 0, 1
        return 1, 0

    Producto.objects.create(
        nombre=data['nombre'],
        descripcion=data['descripcion'],
        categoria=categoria_obj,
        cantidad=data['cantidad'],
        precio_unitario=data['precio_unitario'],
        precio_venta=data['precio_venta'],
        stock_minimo=data['stock_minimo'],
        estado=data['estado'],
        usuario_creador=user,
    )
    return 1, 0

def _procesar_hoja(ws, indice, crear_categorias, user, estados_validos):
    creados, actualizados, errores = 0, 0, []
    for row in ws.iter_rows(min_row=2):
        try:
            data = _normalizar_valores_fila(row, indice, estados_validos)
            c, a = _manejar_producto(data, user, crear_categorias)
            creados += c
            actualizados += a
        except Exception as e:
            errores.append(str(e))
    return creados, actualizados, errores

@login_required
@never_cache
@require_GET
def inventario_dashboard(request):
    """Dashboard principal del inventario"""
    productos = Producto.objects.filter(estado='activo')
    
    # Estadísticas generales
    total_productos = productos.count()
    valor_total_inventario = productos.aggregate(
        total=Sum('cantidad') * Sum('precio_unitario')
    )['total'] or 0
    productos_bajo_stock = productos.filter(cantidad__lte=F('stock_minimo')).count()
    
    # Productos más recientes
    productos_recientes = productos.order_by('-fecha_creacion')[:5]
    
    # Productos con bajo stock
    productos_alerta = productos.filter(cantidad__lte=F('stock_minimo'))[:5]
    
    context = {
        'total_productos': total_productos,
        'valor_total_inventario': valor_total_inventario,
        'productos_bajo_stock': productos_bajo_stock,
        'productos_recientes': productos_recientes,
        'productos_alerta': productos_alerta,
    }
    
    return render(request, 'inventario/dashboard.html', context)

@login_required
@never_cache
@require_GET
def lista_productos(request):
    """Lista todos los productos con filtros y paginación usando utilidades centralizadas"""
    from S_CONTABLE.utils import aplicar_busqueda_texto, paginar_queryset
    
    productos = Producto.objects.all()
    
    # Filtros
    busqueda = request.GET.get('busqueda')
    categoria_id = request.GET.get('categoria')
    estado = request.GET.get('estado')
    
    # Usar helper centralizado para búsqueda de texto
    productos = aplicar_busqueda_texto(productos, busqueda, ['nombre', 'codigo', 'descripcion'])
    
    if categoria_id:
        productos = productos.filter(categoria_id=categoria_id)
    
    if estado:
        productos = productos.filter(estado=estado)
    
    # Usar helper centralizado para paginación
    page_obj = paginar_queryset(productos, request, items_per_page=10)
    
    # Para los filtros
    categorias = Categoria.objects.all()
    
    context = {
        'page_obj': page_obj,
        'categorias': categorias,
        'busqueda': busqueda,
        'categoria_seleccionada': categoria_id,
        'estado_seleccionado': estado,
    }
    
    return render(request, 'inventario/lista_productos.html', context)

@login_required
@never_cache
@require_GET
def detalle_producto(request, producto_id):
    """Muestra el detalle de un producto"""
    producto = get_object_or_404(Producto, id=producto_id)
    movimientos = producto.movimientos.all()[:10]  # Últimos 10 movimientos
    
    context = {
        'producto': producto,
        'movimientos': movimientos,
    }
    
    return render(request, 'inventario/detalle_producto.html', context)

@login_required
@never_cache
# NOSONAR - Django CSRF protection is enabled by default for POST requests
@require_http_methods(['GET', 'POST'])
def crear_producto(request):
    """Crea un nuevo producto"""
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            producto = form.save(commit=False)
            producto.usuario_creador = request.user
            producto.save()
            messages.success(request, f'Producto "{producto.nombre}" creado exitosamente.')
            return redirect(DETALLE_PRODUCTO_URL, producto_id=producto.id)
    else:
        form = ProductoForm()
    
    return render(request, 'inventario/crear_producto.html', {'form': form})

@login_required
@never_cache
# NOSONAR - Django CSRF protection is enabled by default for POST requests
@require_http_methods(['GET', 'POST'])
def editar_producto(request, producto_id):
    """Edita un producto existente"""
    producto = get_object_or_404(Producto, id=producto_id)
    
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            messages.success(request, f'Producto "{producto.nombre}" actualizado exitosamente.')
            return redirect(DETALLE_PRODUCTO_URL, producto_id=producto.id)
    else:
        form = ProductoForm(instance=producto)
    
    return render(request, 'inventario/editar_producto.html', {
        'form': form,
        'producto': producto
    })

@login_required
@never_cache
# NOSONAR - Django CSRF protection is enabled by default for POST requests
@require_http_methods(['GET', 'POST'])
def eliminar_producto(request, producto_id):
    """Elimina (desactiva) un producto"""
    producto = get_object_or_404(Producto, id=producto_id)
    
    if request.method == 'POST':
        producto.estado = 'inactivo'
        producto.save()
        messages.success(request, f'Producto "{producto.nombre}" desactivado exitosamente.')
        return redirect('inventario:lista_productos')
    
    return render(request, 'inventario/confirmar_eliminacion.html', {'producto': producto})

@login_required
@never_cache
# NOSONAR - Django CSRF protection is enabled by default for POST requests
@require_http_methods(['GET', 'POST'])
def crear_movimiento(request, producto_id):
    """Crea un movimiento de inventario"""
    producto = get_object_or_404(Producto, id=producto_id)
    
    if request.method == 'POST':
        form = MovimientoInventarioForm(request.POST)
        if form.is_valid():
            movimiento = form.save(commit=False)
            movimiento.producto = producto
            movimiento.usuario = request.user
            
            # Actualizar cantidad del producto
            if movimiento.tipo == 'entrada':
                producto.cantidad += movimiento.cantidad
            elif movimiento.tipo == 'salida':
                if producto.cantidad >= movimiento.cantidad:
                    producto.cantidad -= movimiento.cantidad
                else:
                    messages.error(request, 'No hay suficiente stock para realizar esta salida.')
                    return render(request, 'inventario/crear_movimiento.html', {
                        'form': form,
                        'producto': producto
                    })
            elif movimiento.tipo == 'ajuste':
                producto.cantidad = movimiento.cantidad
            
            producto.save()
            movimiento.save()
            
            messages.success(request, f'Movimiento registrado exitosamente. Nueva cantidad: {producto.cantidad}')
            return redirect(DETALLE_PRODUCTO_URL, producto_id=producto.id)
    else:
        form = MovimientoInventarioForm()
    
    return render(request, 'inventario/crear_movimiento.html', {
        'form': form,
        'producto': producto
    })

@login_required
@never_cache
@require_GET
def lista_movimientos(request):
    """Lista todos los movimientos de inventario con filtros usando utilidades centralizadas"""
    from S_CONTABLE.utils import aplicar_filtros_fecha, paginar_queryset
    
    movimientos = MovimientoInventario.objects.select_related('producto', 'usuario').all().order_by('-fecha')
    
    # Filtros
    producto_id = request.GET.get('producto')
    tipo = request.GET.get('tipo')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if producto_id:
        movimientos = movimientos.filter(producto_id=producto_id)
    
    if tipo:
        movimientos = movimientos.filter(tipo=tipo)
    
    # Usar helper centralizado para filtros de fecha
    movimientos = aplicar_filtros_fecha(movimientos, fecha_desde, fecha_hasta, campo_fecha='fecha')
    
    # Usar helper centralizado para paginación
    page_obj = paginar_queryset(movimientos, request, items_per_page=20)
    
    # Productos para el filtro
    productos = Producto.objects.filter(estado='activo').order_by('nombre')
    
    context = {
        'page_obj': page_obj,
        'productos': productos,
        'producto_seleccionado': producto_id,
        'tipo_seleccionado': tipo,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }
    
    return render(request, 'inventario/lista_movimientos.html', context)

@login_required
@never_cache
@require_GET
def lista_categorias(request):
    """Lista todas las categorías"""
    categorias = Categoria.objects.all()
    return render(request, 'inventario/lista_categorias.html', {'categorias': categorias})

@login_required
@never_cache
@csrf_protect
@require_http_methods(['GET', 'POST'])
def crear_categoria(request):
    """Crea una nueva categoría"""
    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        if form.is_valid():
            categoria = form.save()
            messages.success(request, f'Categoría "{categoria.nombre}" creada exitosamente.')
            return redirect('inventario:lista_categorias')
    else:
        form = CategoriaForm()
    
    return render(request, 'inventario/crear_categoria.html', {'form': form})

@login_required
@never_cache
@require_GET
def reporte_inventario(request):
    """Genera reporte de inventario"""
    productos = Producto.objects.filter(estado='activo')
    
    # Estadísticas
    total_productos = productos.count()
    valor_total = sum(p.valor_total for p in productos)
    productos_bajo_stock = productos.filter(cantidad__lte=F('stock_minimo'))
    
    context = {
        'productos': productos,
        'total_productos': total_productos,
        'valor_total': valor_total,
        'productos_bajo_stock': productos_bajo_stock,
    }
    
    return render(request, 'inventario/reporte_inventario.html', context)

@login_required
@never_cache
@require_http_methods(['GET', 'POST'])
def importar_productos(request):
    """
    Importa productos desde un archivo Excel (.xlsx) usando helpers para baja complejidad.
    """
    if request.method != 'POST':
        return render(request, TEMPLATE_IMPORTAR_PRODUCTOS, {'form': ImportarProductosForm()})

    form = ImportarProductosForm(request.POST, request.FILES)
    if not form.is_valid():
        return render(request, TEMPLATE_IMPORTAR_PRODUCTOS, {'form': form})

    try:
        contenido = form.cleaned_data['archivo'].read()
        ws = load_workbook(filename=BytesIO(contenido), data_only=True).active

        headers = [str(c.value).strip().lower() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
        indice = {h: i for i, h in enumerate(headers) if h}
        faltantes = [h for h in ['nombre', 'cantidad', 'precio_unitario'] if h not in indice]
        if faltantes:
            messages.error(request, f'Faltan columnas requeridas: {", ".join(faltantes)}')
            return render(request, TEMPLATE_IMPORTAR_PRODUCTOS, {'form': form})

        estados_validos = set(dict(Producto.ESTADO_CHOICES).keys())
        creados, actualizados, errores = _procesar_hoja(
            ws, indice, form.cleaned_data.get('crear_categorias', True), request.user, estados_validos
        )

        if creados or actualizados:
            messages.success(request, f'Importación completada. Creados: {creados}, Actualizados: {actualizados}.')
        if errores:
            messages.warning(request, f'Se encontraron {len(errores)} filas con error.')
        return redirect('inventario:lista_productos')

    except Exception as e:
        messages.error(request, f'No se pudo procesar el archivo: {e}')
        return render(request, TEMPLATE_IMPORTAR_PRODUCTOS, {'form': form})

@login_required
@never_cache
@require_GET
def plantilla_importacion_productos(request):
    """Genera y descarga una plantilla Excel para importar productos."""
    from openpyxl import Workbook
    from openpyxl.writer.excel import save_workbook

    wb = Workbook()
    ws = wb.active
    ws.title = 'Productos'

    headers = [
        'codigo', 'nombre', 'descripcion', 'categoria',
        'cantidad', 'precio_unitario', 'precio_venta', 'stock_minimo', 'estado'
    ]
    ws.append(headers)

    # Fila de ejemplo
    ws.append([
        'PROD0001', 'Camiseta básica', 'Algodón 100%', 'Ropa',
        50, 20000, 30000, 5, 'activo'
    ])

    # Ajuste simple de ancho
    for col in ws.columns:
        max_length = 12
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_importacion_productos.xlsx"'
    save_workbook(wb, response)
    return response